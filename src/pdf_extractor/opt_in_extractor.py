"""Extract telemarketing opt-in data from DealerTrack and RouteOne credit applications.

Strategy per form type and page mode:

  Digital text pages
  ------------------
  RouteOne:   All fields extracted from text.
              - Name from "Credit Application: Applicant" line.
              - Phones from "at the following telephone number(s):".
              - Opt-in from whether a date appears after the "Optional Consent"
                applicant signature line (the SECOND signature on the form).
  DealerTrack: Name and phones extracted from text.
               Opt-in determined by vision only (the checkbox mark is a visual
               element — "You opt in" and "You do not opt in" always appear on
               the same text line regardless of which box is ticked).

  Scanned / image-only pages
  --------------------------
  Both form types: full GPT vision extraction with form-specific prompts.

Required environment variables:
    AZURE_OPENAI_ENDPOINT    e.g. https://my-resource.openai.azure.com/
    AZURE_OPENAI_API_KEY     API key for the Azure OpenAI resource
    AZURE_OPENAI_DEPLOYMENT  Deployment name (default: gpt-4o-mini)

Single-file usage:
    from pypdf import PdfReader
    from pdf_extractor.opt_in_extractor import extract_credit_app_data

    reader = PdfReader("Credit_Application.pdf")
    result = extract_credit_app_data(reader)

Batch usage:
    from pdf_extractor.opt_in_extractor import process_folder_to_excel
    process_folder_to_excel("output/", "results.xlsx")
"""

from __future__ import annotations

import base64
import datetime
import io
import json
import logging
import os
import re
import tempfile
from pathlib import Path

from PIL import Image

_logger = logging.getLogger(__name__)

_DEFAULT_DEPLOYMENT = "gpt-4o-mini"
_CREDIT_APP_FILENAME = "Credit_Application.pdf"
_TEXT_PAGE_MIN_CHARS = 50   # pages with fewer chars are treated as scanned
_DPI_CHECKBOX = 300         # higher DPI for checkbox detection


# ---------------------------------------------------------------------------
# Form-type detection patterns
# ---------------------------------------------------------------------------

_ROUTEONE_RE = re.compile(r"RouteOne", re.IGNORECASE)
_DT_FOOTER_RE = re.compile(r"\bDT\s*\d", re.IGNORECASE)

# ---------------------------------------------------------------------------
# RouteOne text patterns
# ---------------------------------------------------------------------------

# Line immediately after "Credit Application: Applicant" holds the name data.
# Format: [Title] LastName FirstName [Middle] DOB SSN ...
_RO_APPLICANT_RE = re.compile(
    r"Credit Application:\s*Applicant\s*\n(.+)", re.IGNORECASE
)
_RO_TITLE_RE = re.compile(r"^(Mrs?|Miss|Ms|Dr)\.?\s*", re.IGNORECASE)

# Phones listed explicitly in the Optional Consent paragraph.
_RO_PHONES_RE = re.compile(
    r"at the following telephone number\(s\):\s*(.+?)\.",
    re.IGNORECASE | re.DOTALL,
)

# Credit Application signature section: look for the FIRST signature line to check
# if it has a date. If the first (mandatory) signature has no extractable date,
# both signatures are likely handwritten/images, so we should use vision.
_RO_CREDIT_APP_SIG_RE = re.compile(
    r"Credit Application Signature\s*Applicant:\s*By\s+Date(.*?)(?:Optional Consent|$)",
    re.IGNORECASE | re.DOTALL,
)

# Optional Consent section: capture everything between the section header and
# "Source:" (the footer line) so we can look for a date on the consent sig line.
# A blank span means no date was written → opted_out.
_RO_OPT_CONSENT_RE = re.compile(
    r"Optional Consent.*?Applicant:\s*By\s+Date(.*?)(?:Source:|$)",
    re.IGNORECASE | re.DOTALL,
)

# ---------------------------------------------------------------------------
# DealerTrack text patterns
# ---------------------------------------------------------------------------

# Name line: ALL-CAPS LAST FIRST followed by SSN (XXX-XX-XXXX).
_DT_NAME_RE = re.compile(
    r"^([A-Z]{2,})\s+([A-Z]{2,})\s+\d{3}-\d{2}-\d{4}",
    re.MULTILINE,
)

# Customer telemarketing phones on DealerTrack: the blank space between
# "at the following number(s)" and "including any cell phone numbers".
_DT_PHONE_BLOCK_RE = re.compile(
    r"at the following number\(s\)(.*?)(?:including any cell phone numbers|You understand)",
    re.IGNORECASE | re.DOTALL,
)

# Generic phone number pattern (used for extraction from a text block).
_PHONE_RE = re.compile(r"[\+\(]?\d[\d\s\-\(\)\.]{7,}\d")

# Date pattern (numeric MM/DD/YYYY) — used to detect a signed consent line.
_DATE_RE = re.compile(r"\d{1,2}/\d{1,2}/\d{4}")

# Broader date pattern: also matches "January 07, 2025" written-out format.
_VERBOSE_DATE_RE = re.compile(
    r"\d{1,2}/\d{1,2}/\d{4}"
    r"|(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?"
    r"|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)"
    r"\s+\d{1,2},?\s+\d{4}",
    re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# Dealer section patterns (VIN, Dealer ID, Address)
# User note: Dealer # and VIN are on the LAST page of the Credit Application.
# ---------------------------------------------------------------------------

# VIN: any run of 17+ consecutive chars from the VIN character set (no I, O, Q).
# DealerTrack concatenates VIN with model info (e.g. "1C4PJXDG4RW347305SPORT…"),
# so we match 17+ and take the first 17.
_VIN_RUN_RE = re.compile(r"[A-HJ-NPR-Z0-9]{17,}")

# Dealer # label on the same line as the value (no cross-line match).
_DEALER_ID_RE = re.compile(
    r"Dealer\s*(?:#|No\.?|Number)[^\S\n]*:?[^\S\n]*([A-Z0-9]{3,12})",
    re.IGNORECASE,
)

# DealerTrack address: all-caps street number + name on its own line.
# DT forms omit the street type (e.g. "2948 GREENBRIAR" with no "St" suffix).
_DT_STREET_RE = re.compile(r"^(\d+[ \t]+[A-Z][A-Z0-9 \t]{2,40})$", re.MULTILINE)

# DealerTrack city/state/zip appear concatenated: "78156TXSEGUIN"
_DT_ZIP_STATE_CITY_RE = re.compile(
    r"^(\d{5})([A-Z]{2})([A-Z][A-Z\s]+)$", re.MULTILINE
)

# Generic city State ZIP for RouteOne and other forms with proper spacing.
_CITY_STATE_ZIP_RE = re.compile(
    r"([A-Za-z][A-Za-z\s]{1,25}?),?\s+([A-Z]{2})\s+(\d{5}(?:-\d{4})?)"
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _all_text(reader) -> str:
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _is_text_heavy(reader) -> bool:
    """Return True if the PDF is primarily digital text (not a scanned image)."""
    total_chars = sum(len(page.extract_text() or "") for page in reader.pages)
    return total_chars >= _TEXT_PAGE_MIN_CHARS * len(reader.pages)


def _detect_form_type(text: str) -> str:
    if _ROUTEONE_RE.search(text):
        return "routeone"
    if _DT_FOOTER_RE.search(text):
        return "dealertrack"
    return "unknown"


def _extract_vin(last_page_text: str) -> str | None:
    """Extract VIN from the dealer section (last page of credit application).

    Matches any run of 17+ consecutive VIN-charset chars and returns the first 17.
    This handles DealerTrack's concatenated layout (e.g. VIN+model with no separator).
    """
    m = _VIN_RUN_RE.search(last_page_text)
    return m.group(0)[:17] if m else None


def _extract_dealer_id(last_page_text: str) -> str | None:
    """Extract Dealer # from the dealer section (last page of credit application).

    Only matches the value on the same line as the label to avoid false positives
    from adjacent field names on the next line.
    """
    m = _DEALER_ID_RE.search(last_page_text)
    return m.group(1).strip() if m else None


def _extract_dt_address(page_text: str) -> str | None:
    """Extract applicant address from a DealerTrack form page.

    DealerTrack forms have all-caps addresses without street-type suffixes
    (e.g. "2948 GREENBRIAR") and concatenate city/state/zip as "78156TXSEGUIN".
    Searches for both patterns and reconstructs the full address string.
    """
    street_m = _DT_STREET_RE.search(page_text)
    if not street_m:
        return None
    street = street_m.group(1).strip().title()

    csz_m = _DT_ZIP_STATE_CITY_RE.search(page_text)
    if csz_m:
        zip_code = csz_m.group(1)
        state = csz_m.group(2)
        city = csz_m.group(3).strip().title()
        return f"{street}, {city}, {state} {zip_code}"
    return street


def _extract_ro_address(text: str) -> str | None:
    """Best-effort address extraction for RouteOne forms.

    RouteOne forms use normal spacing and street-type suffixes.
    Finds a street pattern then looks nearby for city/state/zip.
    """
    # Street with a type suffix (St, Ave, Blvd, etc.)
    street_re = re.compile(
        r"(\d+\s+(?:[NESW]\.?\s+)?[A-Za-z][A-Za-z0-9\s]{2,30}"
        r"(?:Street|St|Avenue|Ave|Boulevard|Blvd|Road|Rd|Drive|Dr|Lane|Ln"
        r"|Way|Court|Ct|Place|Pl|Highway|Hwy|Parkway|Pkwy|Circle|Cir)"
        r"\.?(?:\s+(?:Apt|Suite|Ste|Unit|#)\s*[\dA-Za-z]+)?)",
        re.IGNORECASE,
    )
    street_m = street_re.search(text)
    if not street_m:
        return None
    street = street_m.group(0).strip()
    nearby = text[street_m.end():street_m.end() + 250]
    csz_m = _CITY_STATE_ZIP_RE.search(nearby)
    if csz_m:
        city = csz_m.group(1).strip().title()
        state = csz_m.group(2)
        zip_code = csz_m.group(3)
        return f"{street}, {city}, {state} {zip_code}"
    return street


def _unique_phones(raw_list: list[str]) -> list[str]:
    """Deduplicate phone numbers by digit content."""
    seen: set[str] = set()
    result: list[str] = []
    for p in raw_list:
        digits = re.sub(r"\D", "", p)
        if len(digits) >= 10 and digits not in seen:
            seen.add(digits)
            result.append(p.strip())
    return result


# ---------------------------------------------------------------------------
# Text-path extractors
# ---------------------------------------------------------------------------

def _routeone_has_first_signature_date(text: str) -> bool:
    """Check if the first (Credit Application) signature has an extractable date.
    
    If the first signature (which is mandatory) has no extractable date, both
    signatures are likely handwritten/images. In that case, vision extraction
    should be used instead of text extraction.
    """
    m = _RO_CREDIT_APP_SIG_RE.search(text)
    if m:
        first_sig_tail = m.group(1).strip()
        return bool(_DATE_RE.search(first_sig_tail))
    return False


def _routeone_from_text(text: str) -> dict:
    """Extract all RouteOne fields from digital text — no vision needed.
    
    Returns a dict with an additional key 'has_first_sig_date' to indicate
    whether the first (mandatory) signature date was found. If False, the
    caller should fall back to vision extraction.
    """
    # --- Name ---
    last_name: str | None = None
    first_name: str | None = None
    m = _RO_APPLICANT_RE.search(text)
    if m:
        line = _RO_TITLE_RE.sub("", m.group(1).strip())
        parts = line.split()
        if len(parts) >= 2:
            last_name, first_name = parts[0], parts[1]

    # --- Phones ---
    phones: list[str] = []
    m = _RO_PHONES_RE.search(text)
    if m:
        raw = [p.strip() for p in re.split(r"[,;]", m.group(1))]
        phones = _unique_phones(raw)

    # --- Opt-in + generated date ---
    # The Optional Consent is the SECOND "Applicant: By Date" block on the form.
    # We capture the text between that line and the "Source:" footer.
    opt_in_status = "opted_out"
    generated_date: str | None = None
    m = _RO_OPT_CONSENT_RE.search(text)
    if m:
        consent_tail = m.group(1).strip()
        date_m = _DATE_RE.search(consent_tail)
        if date_m:
            opt_in_status = "opted_in"
            generated_date = date_m.group(0)

    # Check if first signature has a date (used for vision fallback detection)
    has_first_sig_date = _routeone_has_first_signature_date(text)

    return {
        "last_name": last_name,
        "first_name": first_name,
        "address": _extract_ro_address(text),
        "dealer_id": _extract_dealer_id(text),
        "vin": _extract_vin(text),
        "generated_date": generated_date,
        "opt_in_status": opt_in_status,
        "telemarketing_phones": phones,
        "confidence": "high",
        "has_first_sig_date": has_first_sig_date,  # For fallback detection
    }


def _dealertrack_name_from_text(text: str) -> tuple[str | None, str | None]:
    """Extract LAST, FIRST name from DealerTrack text via SSN anchor pattern."""
    m = _DT_NAME_RE.search(text)
    if m:
        return m.group(1), m.group(2)
    return None, None


def _dealertrack_phones_from_text(text: str) -> list[str]:
    """Extract customer telemarketing phone numbers from DealerTrack text.

    Only looks in the blank space between 'at the following number(s)' and
    'including any cell phone numbers' — where the customer writes in their
    number(s). Returns an empty list when the space is blank.
    """
    m = _DT_PHONE_BLOCK_RE.search(text)
    if not m:
        return []
    block = m.group(1).strip()
    if not block:
        return []
    return _unique_phones(_PHONE_RE.findall(block))


# ---------------------------------------------------------------------------
# Vision-path helpers
# ---------------------------------------------------------------------------

def _render_page(pdf_page, dpi: int = 150) -> Image.Image | None:
    """Render a PDF page to a PIL Image at the given DPI using PyMuPDF."""
    text = (pdf_page.extract_text() or "").strip()
    if len(text) >= _TEXT_PAGE_MIN_CHARS:
        try:
            import fitz
            from pypdf import PdfWriter

            writer = PdfWriter()
            writer.add_page(pdf_page)
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp_path = Path(tmp.name)
                writer.write(tmp)

            doc = fitz.open(str(tmp_path))
            pix = doc[0].get_pixmap(dpi=dpi)
            doc.close()
            tmp_path.unlink(missing_ok=True)
            return Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        except Exception as exc:
            _logger.warning("_render_page: render failed: %s", exc)

    # Scanned page: use the largest embedded image.
    images = getattr(pdf_page, "images", [])
    if images:
        best = max(
            images,
            key=lambda img: (img.image.width * img.image.height) if img.image else 0,
        )
        if best.image:
            return best.image.convert("RGB")
    return None


def _to_b64(image: Image.Image) -> str:
    buf = io.BytesIO()
    image.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode()


def _get_client(model: str | None = None):
    # Allow direct module usage (without cli.py) to pick up local .env changes.
    try:
        from dotenv import load_dotenv

        load_dotenv(override=True)
    except ImportError:
        pass

    endpoint = os.environ.get("AZURE_OPENAI_ENDPOINT")
    api_key = os.environ.get("AZURE_OPENAI_API_KEY")
    if not endpoint or not api_key:
        raise EnvironmentError(
            "AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY must be set."
        )
    try:
        from openai import AzureOpenAI
    except ImportError:
        raise ImportError("openai package is required: pip install openai>=1.0.0")

    deployment = model or os.environ.get("AZURE_OPENAI_DEPLOYMENT", _DEFAULT_DEPLOYMENT)
    client = AzureOpenAI(
        azure_endpoint=endpoint, api_key=api_key, api_version="2024-02-01"
    )
    return client, deployment


_DT_CHECKBOX_PROMPT = """\
This is one page from a DealerTrack credit application.

Near the bottom there are two options side by side:
  "You opt in"     and     "You do not opt in"

Each option has a small box or circle next to it. Look carefully at both boxes.
One of them should have a mark inside it (a checkmark, an X, a dot, or a filled box).

Which box is marked?

Return valid JSON only — no markdown, no explanation:
{
  "opt_in_status": "opted_in" | "opted_out" | "unclear",
  "confidence": "high" | "medium" | "low"
}
"""

_RO_VISION_PROMPT = """\
These are the pages of a RouteOne credit application.

Extract the following and return valid JSON only — no markdown, no explanation:

1. APPLICANT NAME
   Find the "Credit Application: Applicant" section (page 1).
   The next line after that heading has: [Title] LastName FirstName [Middle] ...
   Extract last_name and first_name.

2. OPT-IN STATUS (Optional Consent only)
   There are TWO signature lines on this form.
   The FIRST is labelled "Credit Application Signature" — ignore it.
   The SECOND is labelled "Optional Consent" / "By signing below, you agree to
   receive ... telemarketing messages ...".
   Look ONLY at the Optional Consent applicant signature line.
   If there is a signature or date on it → "opted_in". If blank → "opted_out".

3. TELEMARKETING PHONES
   In the Optional Consent paragraph find "at the following telephone number(s):"
   and list every number that follows.

4. CUSTOMER ADDRESS
   Find the applicant's current/home address in the applicant section.
   Return the full address as a single string (street, city, state, ZIP).

5. DEALER INFORMATION (Dealer Section — usually page 2 or 3)
   Find the "Dealer #" or "Dealer Number" field and extract the dealer ID.
   Find the "VIN" field and extract the 17-character Vehicle Identification Number.

6. GENERATED DATE
   On the Optional Consent page, find the date written next to the applicant
   signature line (the SECOND signature). Return it as MM/DD/YYYY or null.

{
  "last_name": "<string or null>",
  "first_name": "<string or null>",
  "address": "<string or null>",
  "dealer_id": "<string or null>",
  "vin": "<string or null>",
  "generated_date": "<MM/DD/YYYY or null>",
  "opt_in_status": "opted_in" | "opted_out" | "unclear",
  "telemarketing_phones": ["<phone>", ...],
  "confidence": "high" | "medium" | "low"
}
"""

_RO_OPTIN_ONLY_VISION_PROMPT = """\
You are reviewing RouteOne credit application pages to determine the
Optional Consent telemarketing opt-in status and the date on the consent line.

There are TWO signature areas:
1) Credit Application Signature (required)
2) Optional Consent (telemarketing)

IGNORE the first signature entirely.
Look only at the Optional Consent applicant signature/date line (the SECOND
signature area).

Rules:
- If a visible handwritten signature or visible written date appears on the
    Optional Consent line, return opted_in.
- If the Optional Consent line appears blank (no writing), return opted_out.
- If image quality prevents a determination, return unclear.
- Also extract the date written on the Optional Consent line (MM/DD/YYYY),
    or null if blank or unreadable.

Return valid JSON only:
{
    "opt_in_status": "opted_in" | "opted_out" | "unclear",
    "generated_date": "<MM/DD/YYYY or null>",
    "confidence": "high" | "medium" | "low"
}
"""

_DT_VISION_FULL_PROMPT = """\
These are the pages of a DealerTrack credit application (identified by the "DT"
footer on every page).

Extract the following and return valid JSON only — no markdown, no explanation:

1. APPLICANT NAME
   Find the applicant section (section A). The actual filled-in Last Name and
   First Name values (not the field labels).

2. OPT-IN STATUS
   Near the bottom of the form there are two options side by side:
     "You opt in"   and   "You do not opt in"
   Look carefully at the small box next to each option.
   Return "opted_in" if "You opt in" is marked, "opted_out" if "You do not opt in"
   is marked, "unclear" if you cannot tell.

3. TELEMARKETING PHONES
   Find the paragraph about autodialed/telemarketing calls and look for any
   phone numbers the CUSTOMER wrote in (not the dealer's printed phone number).
   Return an empty list if the space is blank.

4. CUSTOMER ADDRESS
   Find the applicant's address in section A.
   Return the full address as a single string (street, city, state, ZIP).

5. DEALER INFORMATION (Dealer Section — usually page 2 or 3)
   Find the "Dealer #" or "Dealer Number" field and extract the dealer ID.
   Find the "VIN" field and extract the 17-character Vehicle Identification Number.

6. GENERATED DATE
   Find the date on the applicant signature line near the opt-in section.
   Return it as MM/DD/YYYY or null if blank.

{
  "last_name": "<string or null>",
  "first_name": "<string or null>",
  "address": "<string or null>",
  "dealer_id": "<string or null>",
  "vin": "<string or null>",
  "generated_date": "<MM/DD/YYYY or null>",
  "opt_in_status": "opted_in" | "opted_out" | "unclear",
  "telemarketing_phones": ["<phone>", ...],
  "confidence": "high" | "medium" | "low"
}
"""


def _call_vision(pages, prompt: str, model: str | None = None,
                 dpi: int = 150) -> dict:
    """Render pages to images and call GPT vision with the given prompt."""
    client, deployment = _get_client(model)

    blocks = []
    for page in pages:
        img = _render_page(page, dpi=dpi)
        if img:
            blocks.append({
                "type": "image_url",
                "image_url": {
                    "url": f"data:image/png;base64,{_to_b64(img)}",
                    "detail": "high",
                },
            })

    if not blocks:
        return {}

    blocks.append({"type": "text", "text": prompt})

    response = client.chat.completions.create(
        model=deployment,
        messages=[{"role": "user", "content": blocks}],
        max_tokens=1024,
        response_format={"type": "json_object"},
    )
    raw = response.choices[0].message.content
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        _logger.warning("_call_vision: non-JSON response: %s", raw)
        return {}


def _find_page_with(reader, search: str) -> int | None:
    """Return the zero-based index of the first page containing search text."""
    for i, page in enumerate(reader.pages):
        if search in (page.extract_text() or ""):
            return i
    return None


# ---------------------------------------------------------------------------
# Main extraction function
# ---------------------------------------------------------------------------

def extract_credit_app_data(pdf_reader, *, model: str | None = None) -> dict:
    """Extract telemarketing opt-in data from a credit application PDF.

    Chooses text-path or vision-path per form type and page content.

    Returns:
        Dict with keys: form_type, last_name, first_name, opt_in_status,
        telemarketing_phones, confidence.
    """
    text = _all_text(pdf_reader)
    form_type = _detect_form_type(text)
    text_heavy = _is_text_heavy(pdf_reader)

    # ------------------------------------------------------------------
    # ROUTEONE — digital text path (with vision fallback for handwritten sigs)
    # ------------------------------------------------------------------
    if form_type == "routeone" and text_heavy:
        data = _routeone_from_text(text)
        # If the first (mandatory) signature has no extractable date, both
        # signatures are likely handwritten/images. Fall back to vision.
        if not data.pop("has_first_sig_date", True):
            _logger.info(
                "extract_credit_app_data: RouteOne first signature has no extractable date; "
                "falling back to vision extraction"
            )
            # Keep name/phones from text and use vision only for opt-in status.
            vision = _call_vision(
                list(pdf_reader.pages), _RO_OPTIN_ONLY_VISION_PROMPT, model=model
            )
            opt_in_status = vision.get("opt_in_status")
            if opt_in_status not in {"opted_in", "opted_out"}:
                opt_in_status = data.get("opt_in_status", "opted_out")
            return {
                "form_type": "routeone",
                "last_name": data.get("last_name"),
                "first_name": data.get("first_name"),
                "address": data.get("address"),
                "dealer_id": data.get("dealer_id"),
                "vin": data.get("vin"),
                "generated_date": vision.get("generated_date"),
                "opt_in_status": opt_in_status,
                "telemarketing_phones": data.get("telemarketing_phones") or [],
                "confidence": vision.get("confidence", "medium"),
            }
        data["form_type"] = "routeone"
        return data

    # ------------------------------------------------------------------
    # ROUTEONE — scanned / image path
    # ------------------------------------------------------------------
    if form_type == "routeone" and not text_heavy:
        vision = _call_vision(list(pdf_reader.pages), _RO_VISION_PROMPT, model=model)
        return {
            "form_type": "routeone",
            "last_name": vision.get("last_name"),
            "first_name": vision.get("first_name"),
            "address": vision.get("address"),
            "dealer_id": vision.get("dealer_id"),
            "vin": vision.get("vin"),
            "generated_date": vision.get("generated_date"),
            "opt_in_status": vision.get("opt_in_status", "unclear"),
            "telemarketing_phones": vision.get("telemarketing_phones") or [],
            "confidence": vision.get("confidence", "low"),
        }

    # ------------------------------------------------------------------
    # DEALERTRACK — digital text path (hybrid: text + targeted vision)
    # ------------------------------------------------------------------
    if form_type == "dealertrack" and text_heavy:
        last_name, first_name = _dealertrack_name_from_text(text)
        phones = _dealertrack_phones_from_text(text)

        # VIN and Dealer # are on the last page (Dealer Section).
        last_page_text = pdf_reader.pages[-1].extract_text() or ""
        vin = _extract_vin(last_page_text)
        dealer_id = _extract_dealer_id(last_page_text)

        # Address is on the applicant page (first page for applicant, page 0).
        applicant_page_text = pdf_reader.pages[0].extract_text() or ""
        address = _extract_dt_address(applicant_page_text)

        # Checkbox state requires vision — render only the page that has it.
        # Generated date: take the LAST date on the opt-in page (the DOB appears
        # earlier in pypdf's reading order; the signature date is at the end).
        generated_date: str | None = None
        opt_in_page_idx = _find_page_with(pdf_reader, "You opt in")
        if opt_in_page_idx is not None:
            page = pdf_reader.pages[opt_in_page_idx]
            opt_in_page_text = page.extract_text() or ""
            all_dates = list(_VERBOSE_DATE_RE.finditer(opt_in_page_text))
            if all_dates:
                generated_date = all_dates[-1].group(0)
            vision = _call_vision([page], _DT_CHECKBOX_PROMPT,
                                  model=model, dpi=_DPI_CHECKBOX)
            opt_in_status = vision.get("opt_in_status", "unclear")
            confidence = vision.get("confidence", "medium")
        else:
            opt_in_status = "not_found"
            confidence = "low"

        return {
            "form_type": "dealertrack",
            "last_name": last_name,
            "first_name": first_name,
            "address": address,
            "dealer_id": dealer_id,
            "vin": vin,
            "generated_date": generated_date,
            "opt_in_status": opt_in_status,
            "telemarketing_phones": phones,
            "confidence": confidence,
        }

    # ------------------------------------------------------------------
    # DEALERTRACK — scanned / image path
    # ------------------------------------------------------------------
    if form_type == "dealertrack" and not text_heavy:
        vision = _call_vision(list(pdf_reader.pages), _DT_VISION_FULL_PROMPT,
                              model=model, dpi=_DPI_CHECKBOX)
        return {
            "form_type": "dealertrack",
            "last_name": vision.get("last_name"),
            "first_name": vision.get("first_name"),
            "address": vision.get("address"),
            "dealer_id": vision.get("dealer_id"),
            "vin": vision.get("vin"),
            "generated_date": vision.get("generated_date"),
            "opt_in_status": vision.get("opt_in_status", "unclear"),
            "telemarketing_phones": vision.get("telemarketing_phones") or [],
            "confidence": vision.get("confidence", "low"),
        }

    # ------------------------------------------------------------------
    # Unknown form type — full vision fallback
    # ------------------------------------------------------------------
    _logger.warning("extract_credit_app_data: unknown form type, falling back to vision")
    vision = _call_vision(list(pdf_reader.pages), _DT_VISION_FULL_PROMPT, model=model)
    return {
        "form_type": "unknown",
        "last_name": vision.get("last_name"),
        "first_name": vision.get("first_name"),
        "address": vision.get("address"),
        "dealer_id": vision.get("dealer_id"),
        "vin": vision.get("vin"),
        "generated_date": vision.get("generated_date"),
        "opt_in_status": vision.get("opt_in_status", "unclear"),
        "telemarketing_phones": vision.get("telemarketing_phones") or [],
        "confidence": vision.get("confidence", "low"),
    }


# ---------------------------------------------------------------------------
# Batch processing
# ---------------------------------------------------------------------------

_MAX_PHONES = 3   # Phone 1 / Phone 2 / Phone 3 columns


def write_results_to_excel(results: list[dict], output_xlsx: str | Path) -> None:
    """Write a list of extraction result dicts to an Excel file.

    Each dict must contain the keys returned by extract_credit_app_data plus
    an optional 'source_file' key used to populate the Source File column.
    Overwrites the file if it already exists.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    output_xlsx = Path(output_xlsx)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opt-In Results"

    phone_headers = [f"Phone {i + 1}" for i in range(_MAX_PHONES)]
    headers = (
        ["Last Name", "First Name", "Address", "Dealer ID", "VIN", "Generated Date"]
        + phone_headers
        + ["Consent", "Source File"]
    )
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for data in results:
        phones = (data.get("telemarketing_phones") or [])[:_MAX_PHONES]
        phone_cells = phones + [""] * (_MAX_PHONES - len(phones))
        ws.append(
            [
                data.get("last_name") or "",
                data.get("first_name") or "",
                data.get("address") or "",
                data.get("dealer_id") or "",
                data.get("vin") or "",
                data.get("generated_date") or "",
            ]
            + phone_cells
            + [data.get("opt_in_status", ""), data.get("source_file", "")]
        )

    wb.save(output_xlsx)
    _logger.info("Wrote %d rows to %s", len(results), output_xlsx)


def append_results_to_excel(results: list[dict], output_xlsx: str | Path) -> None:
    """Append extraction results to an existing Excel file, or create it if absent.

    Loads the workbook when output_xlsx already exists and appends rows without
    repeating the header.  Creates a fresh workbook with a bold header row when
    the file does not exist yet.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    output_xlsx = Path(output_xlsx)
    phone_headers = [f"Phone {i + 1}" for i in range(_MAX_PHONES)]
    headers = (
        ["Last Name", "First Name", "Address", "Dealer ID", "VIN", "Generated Date"]
        + phone_headers
        + ["Consent", "Source File"]
    )

    if output_xlsx.exists():
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Opt-In Results"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    for data in results:
        phones = (data.get("telemarketing_phones") or [])[:_MAX_PHONES]
        phone_cells = phones + [""] * (_MAX_PHONES - len(phones))
        ws.append(
            [
                data.get("last_name") or "",
                data.get("first_name") or "",
                data.get("address") or "",
                data.get("dealer_id") or "",
                data.get("vin") or "",
                data.get("generated_date") or "",
            ]
            + phone_cells
            + [data.get("opt_in_status", ""), data.get("source_file", "")]
        )

    wb.save(output_xlsx)
    _logger.info("Appended %d rows to %s", len(results), output_xlsx)


def process_folder_to_excel(
    input_folder: str | Path,
    output_xlsx: str | Path,
    *,
    model: str | None = None,
) -> int:
    """Find every Credit_Application.pdf under input_folder and write to Excel.

    One row per customer.  Columns: Last Name, First Name, Phone 1–3, Consent,
    Source File.  Overwrites the workbook each run so the sheet stays clean.

    Returns:
        Number of credit applications processed.
    """
    from pypdf import PdfReader

    input_folder = Path(input_folder)
    credit_apps = sorted(input_folder.rglob(_CREDIT_APP_FILENAME))
    if not credit_apps:
        _logger.warning("No %s files found under %s", _CREDIT_APP_FILENAME, input_folder)
        return 0

    results = []
    for pdf_path in credit_apps:
        _logger.info("Processing %s", pdf_path)
        try:
            reader = PdfReader(str(pdf_path))
            data = extract_credit_app_data(reader, model=model)
        except Exception as exc:
            _logger.error("Failed to process %s: %s", pdf_path, exc)
            data = {
                "last_name": None, "first_name": None,
                "address": None, "dealer_id": None, "vin": None,
                "generated_date": None,
                "opt_in_status": "error", "telemarketing_phones": [],
            }
        data["source_file"] = str(pdf_path)
        results.append(data)

    write_results_to_excel(results, output_xlsx)
    return len(results)
